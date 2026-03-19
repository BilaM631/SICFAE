from django.shortcuts import render, get_object_or_404, redirect
from django.views import generic
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.http import HttpResponse, JsonResponse
from django.db.models import Q
from django.urls import reverse, reverse_lazy
import datetime
from django.utils import timezone

from .models import Candidato, PerfilUtilizador, Provincia, Distrito, Vaga, Entrevista
from .forms import (
    FormularioCandidatura, FormularioAutenticacao, FormularioCriacaoUsuario, 
    FormularioValidacaoDocumentos, FormularioCandidaturaManual,
    FormularioCandidaturaEtapa1, FormularioCandidaturaEtapa2,
    EntrevistaForm, AvaliacaoEntrevistaForm
)
from .utils import render_to_pdf, formatar_numero_telefone, despachante_login
from .managers import GestorEstatisticas
from .permissions import (
    obter_candidatos_acessiveis, 
    pode_gerir_candidato,
    pode_ver_candidato,
    obter_exibicao_nivel_usuario,
    obter_perfil_usuario
)

class PaginaInicialView(generic.TemplateView):
    template_name = 'candidaturas/inicio.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['hide_sidebar'] = True
        return context

class ExportarExcelView(LoginRequiredMixin, generic.View):
    def get(self, request, tipo_relatorio='geral', *args, **kwargs):
        import openpyxl
        from openpyxl.utils import get_column_letter

        # 1. Obter dados filtrados (respeita hierarquia)
        queryset_base = obter_candidatos_acessiveis(request.user)
        
        candidatos = queryset_base
        filename_prefix = "candidatos_geral"
        
        if tipo_relatorio == 'pendentes':
            candidatos = queryset_base.filter(estado=Candidato.Estado.PENDENTE)
            filename_prefix = "candidatos_pendentes"
        elif tipo_relatorio == 'rejeitados':
            candidatos = queryset_base.filter(estado=Candidato.Estado.DOCS_REJEITADOS)
            filename_prefix = "candidatos_rejeitados"
        elif tipo_relatorio == 'auditoria':
            # Para auditoria, exportamos o log se for superusuário, ou nada/erro se não for
            if not request.user.is_superuser:
                 return HttpResponse("Apenas superusuários podem exportar auditoria.", status=403)
            # Logica especial para auditoria seria diferente (outro modelo), 
            # mas vamos manter simples por enquanto ou exportar vazio se não implementado a fundo. 
            # O report PDF usa Candidato.history.all(). Vamos tentar fazer igual.
            # Se for complexo misturar, melhor tratar separado. Mas vamos tentar.
            return self.exportar_auditoria(request)

        # 2. Criar Workbook
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={filename_prefix}_{datetime.datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Candidatos"

        # 3. Cabeçalho
        headers = ['Código', 'Nome Completo', 'BI', 'Gênero', 'Vaga', 'Província', 'Distrito', 'Telefone', 'Estado', 'Data Inscrição']
        ws.append(headers)
        
        # Estilo para Cabeçalho
        for col_num, column_title in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = openpyxl.styles.Font(bold=True)
            
        # 4. Linhas
        for c in candidatos:
            row = [
                c.codigo_candidato,
                c.nome_completo,
                c.numero_bi,
                c.get_genero_display(),
                str(c.vaga), # __str__ retorna o titulo
                str(c.provincia),
                str(c.distrito),
                c.numero_telefone,
                c.get_estado_display(),
                c.data_criacao.strftime("%d/%m/%Y") if c.data_criacao else ""
            ]
            ws.append(row)

        wb.save(response)
        return response

    def exportar_auditoria(self, request):
        import openpyxl
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=auditoria_log_{datetime.datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Auditoria Log"
        
        headers = ['Data', 'Usuario', 'Tipo', 'Alteração']
        ws.append(headers)
        
        # history_user vem do django-simple-history
        historico = Candidato.history.all().order_by('-history_date')[:500] 
        
        for h in historico:
            tipo = "Criação" if h.history_type == '+' else "Edição" if h.history_type == '~' else "Remoção"
            user_str = str(h.history_user) if h.history_user else "Sistema"
            row = [
                h.history_date.strftime("%d/%m/%Y %H:%M"),
                user_str,
                tipo,
                f"Estado alterado para {h.estado}" # Simplificação
            ]
            ws.append(row)
            
        wb.save(response)
        return response


class PainelControloView(LoginRequiredMixin, generic.TemplateView):
    template_name = 'candidaturas/painel_controlo.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Inicializar Gestor
        gestor = GestorEstatisticas(self.request.user)
        
        # 1. Contexto de Permissões tudo relacionado da permissões 
        perfil = obter_perfil_usuario(self.request.user)
        eh_central = self.request.user.is_superuser or (perfil and perfil.nivel == PerfilUtilizador.Nivel.CENTRAL)
        eh_provincial = perfil and perfil.nivel == PerfilUtilizador.Nivel.PROVINCIAL
        eh_distrital = perfil and perfil.nivel == PerfilUtilizador.Nivel.DISTRITAL
        
        context['nivel_usuario'] = obter_exibicao_nivel_usuario(self.request.user)
        context['perfil_usuario'] = perfil
        context['eh_central'] = eh_central
        context['eh_provincial'] = eh_provincial
        context['eh_distrital'] = eh_distrital

        # 2. Estatísticas Gerais
        stats_gerais = gestor.obter_estatisticas_gerais()
        context['total_candidatos'] = stats_gerais['total_candidatos']
        context['candidatos_pendentes'] = stats_gerais['candidatos_pendentes']
        context['stats_funcao'] = stats_gerais['stats_funcao']
        context['stats_estado'] = stats_gerais['stats_estado']
        
        # 3. Estatísticas de Admissão
        stats_admissao = gestor.obter_detalhes_admissao()
        context['total_admitidos'] = stats_admissao['total_admitidos']
        context['admitidos_homens'] = stats_admissao['admitidos_homens']
        context['admitidos_mulheres'] = stats_admissao['admitidos_mulheres']

        # 4. Estatísticas Geográficas
        stats_geo = gestor.obter_distribuicao_geografica(eh_central, eh_provincial, perfil)
        context['stats_geo'] = stats_geo # Ensuring the object is available for JS template
        if 'detalhes_provincia' in stats_geo:
            context['detalhes_provincia'] = stats_geo['detalhes_provincia']
        if 'detalhes_distrito' in stats_geo:
            context['detalhes_distrito'] = stats_geo['detalhes_distrito']

        # 5. Lista Recente
        context['candidatos_recentes'] = gestor.obter_candidatos_recentes()

        return context

class ListaVerificacaoView(LoginRequiredMixin, generic.ListView):
    model = Candidato
    template_name = 'candidaturas/lista_candidatos.html'
    context_object_name = 'candidatos'

    def get_queryset(self):
        # Começa com o queryset base que já aplica filtros de permissão (obter_candidatos_acessiveis)
        qs = obter_candidatos_acessiveis(self.request.user)
        
        # Filtros de Pesquisa
        q = self.request.GET.get('q')
        estado = self.request.GET.get('estado')
        filtro_vaga = self.request.GET.get('vaga')
        provincia_id = self.request.GET.get('provincia')
        distrito_id = self.request.GET.get('distrito')
        filtro_genero = self.request.GET.get('genero')

        # Filtragem por Perfil de Utilizador (Hierarquia)
        # A função obter_candidatos_acessiveis já faz a maior parte da filtragem hierárquica.
        # Este bloco pode ser para refinar ou adicionar lógica específica de UI.
        if hasattr(self.request.user, 'perfil') and self.request.user.perfil:
            perfil = self.request.user.perfil
            
            # STAE Central: Pode ver tudo, mas foca em vagas CENTRAIS para aprovação
            if perfil.nivel == PerfilUtilizador.Nivel.CENTRAL:
                # Opcional: Pode filtrar por vagas centrais se desejar aprovar apenas as suas
                # Mas geralmente central pode ver tudo.
                pass
                
            # STAE Provincial: Vê candidatos da sua província 
            # E Vagas com aprovação PROVINCIAL ou DISTRITAL (para supervisão)
            elif perfil.nivel == PerfilUtilizador.Nivel.PROVINCIAL:
                if perfil.provincia:
                    # A função obter_candidatos_acessiveis já filtra por província para este nível
                    # qs = qs.filter(provincia=perfil.provincia)
                    pass
                    
            # STAE Distrital: Vê apenas candidatos do seu distrito
            # E apenas vagas com aprovação DISTRITAL
            elif perfil.nivel == PerfilUtilizador.Nivel.DISTRITAL:
                if perfil.distrito:
                    # A função obter_candidatos_acessiveis já filtra por distrito para este nível
                    # qs = qs.filter(distrito=perfil.distrito)
                    # Filtra também para mostrar apenas vagas que o distrito pode aprovar
                    # qs = qs.filter(vaga__nivel_aprovacao=Vaga.NivelAprovacao.DISTRITAL) 
                    # Decisão: Distrital vê candidatos locais de qualquer vaga, ou só das que aprova?
                    # Geralmente vê todos locais, mas só tem ação nas que aprova.
                    pass
                    
        # Aplicação dos Filtros de Pesquisa UI
        if q:
            qs = qs.filter(
                Q(nome_completo__icontains=q) | 
                Q(numero_bi__icontains=q) |
                Q(codigo_candidato__icontains=q)
            )
            
        if estado:
            if estado == 'TODOS': # Handle 'TODOS' option
                pass
            else:
                qs = qs.filter(estado=estado)
        if filtro_vaga:
            qs = qs.filter(vaga__id=filtro_vaga)
            
        if filtro_genero:
            qs = qs.filter(genero=filtro_genero)

        # Filtro de Província (Apenas Central/Superuser)
        filtro_provincia = self.request.GET.get('provincia')
        if (self.request.user.is_superuser or obter_perfil_usuario(self.request.user).nivel == PerfilUtilizador.Nivel.CENTRAL) and filtro_provincia:
             qs = qs.filter(provincia__id=filtro_provincia)

        return qs.order_by('-data_criacao')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['estado_atual'] = self.request.GET.get('estado', Candidato.Estado.PENDENTE)
        context['vaga_atual'] = self.request.GET.get('vaga', '')
        context['genero_atual'] = self.request.GET.get('genero', '')
        
        from .models import Vaga, Provincia
        context['vagas'] = Vaga.objects.filter(ativa=True)
        context['generos'] = Candidato.Genero.choices
        
        ul = obter_exibicao_nivel_usuario(self.request.user)
        context['nivel_usuario'] = ul
        context['apenas_visualizacao'] = "Superusuário" in ul or "STAE Central" in ul
        
        # Provincia context for Central/Superuser
        if "Superusuário" in ul or "STAE Central" in ul:
            context['provincias'] = Provincia.objects.all().order_by('nome')
            context['provincia_atual'] = self.request.GET.get('provincia', '')
            
        return context

class DetalheCandidatoView(LoginRequiredMixin, generic.DetailView):
    model = Candidato
    template_name = 'candidaturas/detalhe_candidato.html'
    context_object_name = 'candidato'
    
    def get_object(self, queryset=None):
        obj = super().get_object(queryset)
        if not pode_ver_candidato(self.request.user, obj):
            raise PermissionDenied("Não tem permissão para aceder a este candidato.")
        return obj
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['nivel_usuario'] = obter_exibicao_nivel_usuario(self.request.user)
        context['pode_gerir'] = pode_gerir_candidato(self.request.user, self.object)
        return context



def gerar_pdf(request, pk):
    candidato = get_object_or_404(Candidato, pk=pk)
    data = { 
        'candidato': candidato,
        'data': datetime.datetime.now(),
        'cabecalho_stae': 'SECRETARIADO TÉCNICO DE ADMINISTRAÇÃO ELEITORAL',
    }
    pdf = render_to_pdf('candidaturas/formulario_stae.html', data)
    if pdf:
        response = pdf
        filename = f"Ficha_STAE_{candidato.numero_bi}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    return HttpResponse("Erro ao gerar PDF")


def gerar_ficha_entrevista(request, pk):
    candidato = get_object_or_404(Candidato, pk=pk)
    data = { 
        'candidato': candidato,
        'data': datetime.datetime.now(),
    }
    pdf = render_to_pdf('candidaturas/pdf/ficha_entrevista.html', data)
    if pdf:
        response = pdf
        filename = f"Ficha_Entrevista_{candidato.nome_completo.replace(' ', '_')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    return HttpResponse("Erro ao gerar PDF de Entrevista")

# AJAX Views
def carregar_distritos(request):
    provincia_id = request.GET.get('provincia')
    distritos = Distrito.objects.filter(provincia_id=provincia_id).order_by('nome')
    return JsonResponse(list(distritos.values('id', 'nome')), safe=False)


class RegistarCandidaturaView(generic.CreateView):
    model = Candidato
    form_class = FormularioCandidatura
    template_name = 'candidaturas/ficha_candidatura.html'
    success_url = reverse_lazy('candidaturas:sucesso_candidatura')

    def form_valid(self, form):
        form.instance.estado = Candidato.Estado.PENDENTE
        return super().form_valid(form)

def sucesso_candidatura(request):
    return render(request, 'candidaturas/sucesso_candidatura.html')

# --- Registo em Duas Etapas ---

class RegistarCandidaturaEtapa1View(generic.FormView):
    """Etapa 1: Recolha de Dados Pessoais"""
    form_class = FormularioCandidaturaEtapa1
    template_name = 'candidaturas/candidatura_etapa1.html'
    
    def get_initial(self):
        """Preencher com dados da sessão se existirem (voltar da etapa 2)"""
        initial = super().get_initial()
        dados_sessao = self.request.session.get('candidatura_etapa1', {})
        if dados_sessao:
            initial.update(dados_sessao)
        return initial
    
    def form_valid(self, form):
        """Guardar dados na sessão e redirecionar para Etapa 2"""
        dados = {
            'vaga': form.cleaned_data['vaga'].id,
            'nome_completo': form.cleaned_data['nome_completo'],
            'genero': form.cleaned_data['genero'],
            'data_nascimento': form.cleaned_data['data_nascimento'].isoformat() if form.cleaned_data.get('data_nascimento') else None,
            'numero_bi': form.cleaned_data['numero_bi'],
            'numero_telefone': form.cleaned_data['numero_telefone'],
            'provincia': form.cleaned_data['provincia'].id,
            'distrito': form.cleaned_data['distrito'].id,
            'endereco': form.cleaned_data['endereco'],
        }
        self.request.session['candidatura_etapa1'] = dados
        return redirect('candidaturas:registar_etapa2')

class RegistarCandidaturaEtapa2View(generic.FormView):
    """Etapa 2: Upload de Documentos"""
    form_class = FormularioCandidaturaEtapa2
    template_name = 'candidaturas/candidatura_etapa2.html'
    success_url = reverse_lazy('candidaturas:sucesso_candidatura')
    
    def dispatch(self, request, *args, **kwargs):
        """Verificar se a Etapa 1 foi completada"""
        if 'candidatura_etapa1' not in request.session:
            messages.warning(request, "Por favor, preencha primeiro os seus dados pessoais.")
            return redirect('candidaturas:registar_etapa1')
        return super().dispatch(request, *args, **kwargs)
    
    def get_context_data(self, **kwargs):
        """Adicionar dados da Etapa 1 ao contexto para exibição"""
        context = super().get_context_data(**kwargs)
        dados_etapa1 = self.request.session.get('candidatura_etapa1', {})
        
        # Carregar objetos relacionados para exibição
        from .models import Vaga
        if dados_etapa1:
            context['dados_etapa1'] = {
                'vaga': Vaga.objects.filter(id=dados_etapa1.get('vaga')).first(),
                'nome_completo': dados_etapa1.get('nome_completo'),
                'genero': dict(Candidato.Genero.choices).get(dados_etapa1.get('genero')),
                'data_nascimento': dados_etapa1.get('data_nascimento'),
                'numero_bi': dados_etapa1.get('numero_bi'),
                'numero_telefone': dados_etapa1.get('numero_telefone'),
                'provincia': Provincia.objects.filter(id=dados_etapa1.get('provincia')).first(),
                'distrito': Distrito.objects.filter(id=dados_etapa1.get('distrito')).first(),
                'endereco': dados_etapa1.get('endereco'),
            }
        return context
    
    def form_valid(self, form):
        """Combinar dados da sessão com arquivos e criar candidato"""
        dados_etapa1 = self.request.session.get('candidatura_etapa1')
        
        # Criar instância do candidato
        from datetime import datetime
        candidato = Candidato()
        
        # Dados da Etapa 1
        from .models import Vaga
        candidato.vaga_id = dados_etapa1['vaga']
        candidato.nome_completo = dados_etapa1['nome_completo']
        candidato.genero = dados_etapa1['genero']
        if dados_etapa1.get('data_nascimento'):
            candidato.data_nascimento = datetime.fromisoformat(dados_etapa1['data_nascimento']).date()
        candidato.numero_bi = dados_etapa1['numero_bi']
        candidato.numero_telefone = dados_etapa1['numero_telefone']
        candidato.provincia_id = dados_etapa1['provincia']
        candidato.distrito_id = dados_etapa1['distrito']
        candidato.endereco = dados_etapa1['endereco']
        
        # Dados da Etapa 2 (arquivos)
        candidato.arquivo_cv = form.cleaned_data.get('arquivo_cv')
        candidato.arquivo_bi = form.cleaned_data.get('arquivo_bi')
        candidato.arquivo_certificado = form.cleaned_data.get('arquivo_certificado')
        candidato.foto = form.cleaned_data.get('foto')
        
        # Estado inicial
        candidato.estado = Candidato.Estado.PENDENTE
        
        # Salvar
        candidato.save()
        
        # Limpar sessão
        del self.request.session['candidatura_etapa1']
        
        messages.success(self.request, f"Candidatura submetida com sucesso! Código: {candidato.codigo_candidato} | BI: {candidato.numero_bi}")
        return super().form_valid(form)

def voltar_etapa1(request):
    """Permite voltar à Etapa 1 mantendo os dados"""
    return redirect('candidaturas:registar_etapa1')


def enviar_sms(request, pk):
    candidato = get_object_or_404(Candidato, pk=pk)
    
    telefone = formatar_numero_telefone(candidato.numero_telefone)
    msg = f"Sr(a) {candidato.nome_completo}, foi aprovado para a fase de entrevistas. Compareça no local X as 7h30."
    
    if candidato.estado == Candidato.Estado.PENDENTE:
        candidato.estado = Candidato.Estado.ENTREVISTA_AGENDADA
        candidato.save()
        messages.success(request, "Entrevista agendada com sucesso. A direcionar para o WhatsApp...")
        
    whatsapp_url = f"https://wa.me/{telefone}?text={msg}"
    return redirect(whatsapp_url)

def registar_entrevista(request, pk, resultado):
    candidato = get_object_or_404(Candidato, pk=pk)
    if candidato.estado == Candidato.Estado.ENTREVISTA_AGENDADA:
        if resultado == 'passou':
            candidato.estado = Candidato.Estado.ENTREVISTA_APROVADA
            messages.success(request, "Candidato Aprovado na Entrevista.")
        else:
            candidato.estado = Candidato.Estado.ENTREVISTA_REPROVADA
            messages.warning(request, "Candidato Reprovado na Entrevista.")
        candidato.save()
    return redirect('candidaturas:detalhe_candidato', pk=pk)

def enviar_para_formacao(request, pk):
    candidato = get_object_or_404(Candidato, pk=pk)
    if candidato.estado == Candidato.Estado.ENTREVISTA_APROVADA:
        candidato.estado = Candidato.Estado.EM_FORMACAO
        candidato.save()
        messages.success(request, f"Candidato {candidato.nome_completo} enviado para Formação.")
    return redirect('candidaturas:detalhe_candidato', pk=pk)



# --- Painel do Candidato & Auth ---

class LoginCandidatoView(generic.FormView):
    template_name = 'candidaturas/candidato_login.html'
    form_class = FormularioAutenticacao
    success_url = reverse_lazy('candidaturas:painel_candidato')

    def form_valid(self, form):
        bi = form.cleaned_data['numero_bi']
        telefone = form.cleaned_data['numero_telefone']
        
        try:
            candidato = Candidato.objects.get(numero_bi=bi, numero_telefone=telefone)
            self.request.session['candidato_id'] = candidato.pk
            return super().form_valid(form)
        except Candidato.DoesNotExist:
            form.add_error(None, "Dados não encontrados. Verifique o BI e Telefone.")
            return self.form_invalid(form)

class PainelCandidatoView(generic.TemplateView):
    template_name = 'candidaturas/candidato_dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        candidato = None
        candidato_id = self.request.session.get('candidato_id')
        
        if candidato_id:
            candidato = Candidato.objects.filter(pk=candidato_id).first()
            
        if not candidato and self.request.user.is_authenticated:
            candidato = Candidato.objects.filter(numero_bi=self.request.user.username).first()

        context['candidato'] = candidato
        return context

    def get(self, request, *args, **kwargs):
        candidato_id = request.session.get('candidato_id')
        usuario_eh_candidato = False
        
        if request.user.is_authenticated:
             if Candidato.objects.filter(numero_bi=request.user.username).exists():
                 usuario_eh_candidato = True
        
        if not candidato_id and not usuario_eh_candidato:
            return redirect('candidaturas:login_candidato')
            
        return super().get(request, *args, **kwargs)

def logout_candidato(request):
    if 'candidato_id' in request.session:
        del request.session['candidato_id']
    return redirect('candidaturas:inicio')

# Dispatcher
def despachante_auth(request):
    return redirect(despachante_login(request.user))

class ImportarExcelView(LoginRequiredMixin, generic.TemplateView):
    template_name = 'candidaturas/importar_excel.html'

    def post(self, request, *args, **kwargs):
        if 'excel_file' not in request.FILES:
            messages.error(request, "Nenhum arquivo enviado.")
            return redirect('.')
        
        excel_file = request.FILES['excel_file']
        import openpyxl
        
        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            
            atualizados = 0
            criados = 0
            
            rows = list(ws.rows)
            header = [c.value for c in rows[0]] if rows else []
            
            def find_col(names):
                for i, h in enumerate(header):
                    if str(h).strip().lower() in [n.lower() for n in names]:
                        return i
                return -1

            bi_idx = find_col(['BI', 'Número de BI', 'Numero de BI'])
            obs_idx = find_col(['Observacoes', 'Observações', 'Notas', 'Obs'])
            nome_idx = find_col(['Nome', 'Nome Completo'])
            tel_idx = find_col(['Telefone', 'Telemovel', 'Celular'])
            vaga_idx = find_col(['Vaga', 'Cargo', 'Função'])
            
            if bi_idx == -1:
                messages.error(request, "Coluna 'BI' não encontrada no Excel.")
                return redirect('.')

            for row in rows[1:]:
                bi_val = row[bi_idx].value
                if not bi_val: continue
                
                bi_val = str(bi_val).strip()
                obs_val = row[obs_idx].value if obs_idx != -1 else ""
                nome_val = row[nome_idx].value if nome_idx != -1 else ""
                tel_val = row[tel_idx].value if tel_idx != -1 else ""
                vaga_val = row[vaga_idx].value if vaga_idx != -1 else ""
                
                if obs_val is None: obs_val = ""
                
                funcao_valida = Candidato.Funcao.BRIGADISTA
                if vaga_val:
                    r_str = str(vaga_val).strip().upper()
                    if 'FORMADOR' in r_str: funcao_valida = Candidato.Funcao.FORMADOR
                    elif 'BRIGADISTA' in r_str: funcao_valida = Candidato.Funcao.BRIGADISTA
                    elif 'AGENTE' in r_str or 'CIVICO' in r_str: funcao_valida = Candidato.Funcao.AGENTE_CIVICO

                candidato, criado = Candidato.objects.get_or_create(
                    numero_bi=bi_val,
                    defaults={
                        'nome_completo': nome_val or f"Candidato {bi_val}",
                        'numero_telefone': tel_val or "000000000",
                        'observacoes': obs_val,
                        'funcao': funcao_valida,
                        'estado': Candidato.Estado.PENDENTE
                    }
                )

                if not criado:
                    updates = []
                    if obs_val:
                        if candidato.observacoes:
                            candidato.observacoes += f"\n[Import]: {obs_val}"
                        else:
                            candidato.observacoes = obs_val
                        updates.append("obs")
                    if updates:
                        candidato.save()
                        atualizados += 1
                else:
                    criados += 1

            messages.success(request, f"Importação concluída. Atualizados: {atualizados}, Criados: {criados}")
            
        except Exception as e:
            messages.error(request, f"Erro ao processar arquivo: {str(e)}")
            
        return redirect('.')         

class GerirUtilizadoresView(LoginRequiredMixin, generic.CreateView):
    template_name = 'candidaturas/gestao_utilizadores.html'
    form_class = FormularioCriacaoUsuario
    success_url = reverse_lazy('candidaturas:gestao_utilizadores')

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return self.handle_no_permission()
            
        perfil = obter_perfil_usuario(self.request.user)
        if not request.user.is_superuser:
             if not perfil:
                 raise PermissionDenied("Não tem permissão para gerir utilizadores.")
        return super().dispatch(request, *args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        perfil = obter_perfil_usuario(self.request.user)
        
        if self.request.user.is_superuser or (perfil and perfil.nivel == PerfilUtilizador.Nivel.CENTRAL):
             if self.request.user.is_superuser:
                 context['lista_usuarios'] = PerfilUtilizador.objects.filter(
                     Q(nivel=PerfilUtilizador.Nivel.PROVINCIAL) | Q(nivel=PerfilUtilizador.Nivel.CENTRAL)
                 ).select_related('usuario', 'provincia')
             else:
                 context['lista_usuarios'] = PerfilUtilizador.objects.filter(nivel=PerfilUtilizador.Nivel.PROVINCIAL).select_related('usuario', 'provincia')
             context['titulo_pagina'] = "Gestão de Utilizadores Provinciais"
        elif perfil and perfil.nivel == PerfilUtilizador.Nivel.PROVINCIAL:
             context['lista_usuarios'] = PerfilUtilizador.objects.filter(
                 nivel=PerfilUtilizador.Nivel.DISTRITAL, 
                 provincia=perfil.provincia
             ).select_related('usuario', 'distrito')
             context['titulo_pagina'] = f"Gestão de Utilizadores Distritais - {perfil.provincia.nome}"
        elif perfil and perfil.nivel == PerfilUtilizador.Nivel.DISTRITAL:
             context['lista_usuarios'] = PerfilUtilizador.objects.filter(
                 distrito=perfil.distrito
             ).exclude(usuario=self.request.user).select_related('usuario', 'distrito')
             context['titulo_pagina'] = f"Gestão de Utilizadores - {perfil.distrito.nome}"
             
        context['nivel_usuario'] = obter_exibicao_nivel_usuario(self.request.user)
        return context
        
    def form_valid(self, form):
        messages.success(self.request, "Utilizador criado com sucesso.")
        return super().form_valid(form)



class RelatoriosView(LoginRequiredMixin, generic.TemplateView):
    template_name = 'candidaturas/relatorios.html'

def relatorio_pdf(request, tipo_relatorio):
    user = request.user
    if not user.is_authenticated:
        return redirect('login')
        
    queryset_base = obter_candidatos_acessiveis(user).select_related('provincia', 'distrito').order_by('nome_completo')
    
    titulo = "Relatório"
    template = 'candidaturas/pdf/lista_generica.html'
    context = {}
    
    if tipo_relatorio == 'geral':
        titulo = "Lista Geral de Inscritos"
        queryset = queryset_base
        
    elif tipo_relatorio == 'pendentes':
        titulo = "Relatório de Candidaturas Pendentes"
        queryset = queryset_base.filter(estado=Candidato.Estado.PENDENTE)
        
    elif tipo_relatorio == 'rejeitados':
        titulo = "Relatório de Candidaturas Rejeitadas"
        queryset = queryset_base.filter(estado=Candidato.Estado.DOCS_REJEITADOS)
        
    elif tipo_relatorio == 'estatisticas':
        titulo = "Resumo Estatístico de Candidaturas"
        gestor = GestorEstatisticas(user)
        stats_gerais = gestor.obter_estatisticas_gerais()
        
        template = 'candidaturas/pdf/resumo_estatistico.html'
        context['stats_total'] = stats_gerais['total_candidatos']
        context['stats_role'] = stats_gerais['stats_funcao']
        context['stats_status'] = stats_gerais['stats_estado']
        
        queryset = queryset_base

    elif tipo_relatorio == 'auditoria':
        if not user.is_superuser:
             raise PermissionDenied("Apenas superusuários podem gerar relatórios de auditoria.")
             
        titulo = "Relatório de Auditoria do Sistema (Audit Log)"
        template = 'candidaturas/pdf/auditoria_log.html'
        
        # Obter histórico (últimos 100 eventos)
        # Note: Candidato.history.all() vem do django-simple-history
        historico = Candidato.history.all().order_by('-history_date')[:100]
        context['historico'] = historico
        queryset = Candidato.objects.none() # Não usado neste template

    else:
        titulo = "Relatório Desconhecido"
        queryset = Candidato.objects.none()

    if tipo_relatorio != 'estatisticas':
        context['candidatos'] = queryset

    context['titulo'] = titulo
    context['area_nome'] = obter_exibicao_nivel_usuario(user)
    context['data'] = datetime.datetime.now()

    pdf = render_to_pdf(template, context)
    if pdf:
        response = pdf
        filename = f"Relatorio_{tipo_relatorio}_{datetime.datetime.now().strftime('%Y%m%d')}.pdf"
        response['Content-Disposition'] = f'inline; filename="{filename}"'
        return response
    
    return HttpResponse("Erro ao gerar PDF")

class PesquisaCandidatoView(LoginRequiredMixin, generic.View):
    template_name = 'candidaturas/pesquisa_candidato.html'

    def get(self, request):
        return render(request, self.template_name)

    def post(self, request):
        bi = request.POST.get('numero_bi', '').strip()
        if not bi:
            messages.error(request, "Por favor, introduza um número de BI.")
            return render(request, self.template_name)
        
        try:
            candidato = Candidato.objects.get(numero_bi__iexact=bi)
            
            if not pode_ver_candidato(request.user, candidato):
                 messages.error(request, "Não tem permissão para aceder a este candidato.")
                 return render(request, self.template_name)

            url = reverse_lazy('candidaturas:detalhe_candidato', kwargs={'pk': candidato.pk})
            return redirect(f"{url}?tab=docs")
            
        except Candidato.DoesNotExist:
            messages.error(request, f"Candidato com BI '{bi}' não encontrado.")
            return render(request, self.template_name)

class RegistarCandidaturaManualView(LoginRequiredMixin, generic.CreateView):
    model = Candidato
    form_class = FormularioCandidaturaManual
    template_name = 'candidaturas/registar_manual.html'
    
    def get_success_url(self):
        messages.success(self.request, "Candidato registado manualmente e documentos validados.")
        return reverse_lazy('candidaturas:painel_controlo')

    def form_valid(self, form):
        candidato = form.save(commit=False)
        candidato.estado = Candidato.Estado.PENDENTE
             
        candidato.save()
        return redirect(self.get_success_url())

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from .models import Provincia
        context['provincias'] = Provincia.objects.all()
        return context

def consulta_publica(request):
    candidato = None
    erro = None
    termo = None
    
    if request.method == 'POST':
        termo = request.POST.get('termo', '').strip()
        if termo:
            try:
                # Busca por BI ou Telefone
                candidato = Candidato.objects.filter(
                    Q(numero_bi__iexact=termo) | Q(numero_telefone=termo)
                ).first()
                
                if not candidato:
                    erro = "Nenhuma candidatura encontrada com os dados fornecidos."
            except Exception as e:
                erro = "Ocorreu um erro na consulta."
        else:
            erro = "Por favor, introduza um número de BI ou telefone."
            
    return render(request, 'candidaturas/consulta_publica.html', {'candidato': candidato, 'erro': erro, 'termo': termo})


# ========== GEST�O DE VAGAS ==========

# ========== GESTÃO DE VAGAS ==========

class ListaVagasView(LoginRequiredMixin, generic.ListView):
    """Lista todas as vagas com filtros de status."""
    model = Vaga
    template_name = 'candidaturas/vagas/lista_vagas.html'
    context_object_name = 'vagas'
    paginate_by = 10
    
    def get_queryset(self):
        queryset = Vaga.objects.all().order_by('-data_criacao')
        
        # Filtro de status
        status = self.request.GET.get('status')
        if status == 'ativas':
            queryset = queryset.filter(ativa=True)
        elif status == 'inativas':
            queryset = queryset.filter(ativa=False)
        
        # Filtro de pesquisa
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(titulo__icontains=search) | Q(descricao__icontains=search)
            )
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['status_atual'] = self.request.GET.get('status', 'todas')
        context['search_query'] = self.request.GET.get('search', '')
        return context


class CriarVagaView(LoginRequiredMixin, generic.CreateView):
    """Cria uma nova vaga."""
    model = Vaga
    template_name = 'candidaturas/vagas/form_vaga.html'
    fields = ['titulo', 'descricao', 'data_inicio', 'data_fim', 'ativa']
    success_url = reverse_lazy('candidaturas:lista_vagas')
    
    def form_valid(self, form):
        messages.success(self.request, f'Vaga "{form.instance.titulo}" criada com sucesso!')
        return super().form_valid(form)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = 'Nova Vaga'
        context['botao_texto'] = 'Criar Vaga'
        return context


class EditarVagaView(LoginRequiredMixin, generic.UpdateView):
    """Edita uma vaga existente."""
    model = Vaga
    template_name = 'candidaturas/vagas/form_vaga.html'
    fields = ['titulo', 'descricao', 'data_inicio', 'data_fim', 'ativa']
    success_url = reverse_lazy('candidaturas:lista_vagas')
    
    def form_valid(self, form):
        messages.success(self.request, f'Vaga "{form.instance.titulo}" atualizada com sucesso!')
        return super().form_valid(form)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = 'Editar Vaga'
        context['botao_texto'] = 'Salvar Alterações'
        return context


class DetalheVagaView(LoginRequiredMixin, generic.DetailView):
    """Exibe detalhes de uma vaga com estatísticas."""
    model = Vaga
    template_name = 'candidaturas/vagas/detalhe_vaga.html'
    context_object_name = 'vaga'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        vaga = self.object
        
        # Estatísticas da vaga
        context['total_candidatos'] = vaga.candidatos.count()
        context['candidatos_pendentes'] = vaga.candidatos.filter(estado=Candidato.Estado.PENDENTE).count()
        context['candidatos_docs_aprovados'] = vaga.candidatos.filter(estado=Candidato.Estado.DOCS_APROVADOS).count()
        context['candidatos_entrevista_agendada'] = vaga.candidatos.filter(estado=Candidato.Estado.ENTREVISTA_AGENDADA).count()
        context['candidatos_aprovados'] = vaga.candidatos.filter(estado=Candidato.Estado.ENTREVISTA_APROVADA).count()
        context['candidatos_enviados_defc'] = vaga.candidatos.filter(estado=Candidato.Estado.ENVIADO_DEFC).count()
        
        # Lista de candidatos recentes
        context['candidatos_recentes'] = vaga.candidatos.all().order_by('-data_criacao')[:10]
        
        return context


def alternar_status_vaga(request, pk):
    """Ativa ou desativa uma vaga."""
    vaga = get_object_or_404(Vaga, pk=pk)
    vaga.ativa = not vaga.ativa
    vaga.save()
    
    status = "ativada" if vaga.ativa else "desativada"
    messages.success(request, f'Vaga "{vaga.titulo}" {status} com sucesso!')
    
    return redirect('candidaturas:lista_vagas')


# ========== GEST�O DE VAGAS ==========

class ListaVagasView(LoginRequiredMixin, generic.ListView):
    model = Vaga
    template_name = 'candidaturas/vagas/lista_vagas.html'
    context_object_name = 'vagas'
    paginate_by = 10
    
    def get_queryset(self):
        queryset = Vaga.objects.all().order_by('-data_criacao')
        status = self.request.GET.get('status')
        if status == 'ativas':
            queryset = queryset.filter(ativa=True)
        elif status == 'inativas':
            queryset = queryset.filter(ativa=False)
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(Q(titulo__icontains=search) | Q(descricao__icontains=search))
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['status_atual'] = self.request.GET.get('status', 'todas')
        context['search_query'] = self.request.GET.get('search', '')
        return context


class CriarVagaView(LoginRequiredMixin, generic.CreateView):
    model = Vaga
    template_name = 'candidaturas/vagas/form_vaga.html'
    fields = ['titulo', 'descricao', 'data_inicio', 'data_fim', 'ativa']
    success_url = reverse_lazy('candidaturas:lista_vagas')
    
    def form_valid(self, form):
        messages.success(self.request, f'Vaga "{form.instance.titulo}" criada com sucesso!')
        return super().form_valid(form)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = 'Nova Vaga'
        context['botao_texto'] = 'Criar Vaga'
        return context


class EditarVagaView(LoginRequiredMixin, generic.UpdateView):
    model = Vaga
    template_name = 'candidaturas/vagas/form_vaga.html'
    fields = ['titulo', 'descricao', 'data_inicio', 'data_fim', 'ativa']
    success_url = reverse_lazy('candidaturas:lista_vagas')
    
    def form_valid(self, form):
        messages.success(self.request, f'Vaga "{form.instance.titulo}" atualizada com sucesso!')
        return super().form_valid(form)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = 'Editar Vaga'
        context['botao_texto'] = 'Salvar Altera��es'
        return context


class DetalheVagaView(LoginRequiredMixin, generic.DetailView):
    model = Vaga
    template_name = 'candidaturas/vagas/detalhe_vaga.html'
    context_object_name = 'vaga'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        vaga = self.object
        context['total_candidatos'] = vaga.candidatos.count()
        context['candidatos_pendentes'] = vaga.candidatos.filter(estado=Candidato.Estado.PENDENTE).count()
        context['candidatos_entrevista_agendada'] = vaga.candidatos.filter(estado=Candidato.Estado.ENTREVISTA_AGENDADA).count()
        context['candidatos_aprovados'] = vaga.candidatos.filter(estado=Candidato.Estado.ENTREVISTA_APROVADA).count()
        context['candidatos_enviados_defc'] = vaga.candidatos.filter(estado=Candidato.Estado.ENVIADO_DEFC).count()
        context['candidatos_recentes'] = vaga.candidatos.all().order_by('-data_criacao')[:10]
        return context


def alternar_status_vaga(request, pk):
    vaga = get_object_or_404(Vaga, pk=pk)
    vaga.ativa = not vaga.ativa
    vaga.save()
    status = "ativada" if vaga.ativa else "desativada"
    messages.success(request, f'Vaga "{vaga.titulo}" {status} com sucesso!')
    return redirect('candidaturas:lista_vagas')


# Importar views de vagas
from .views_vagas import ListaVagasView, EditarVagaView, DetalheVagaView, alternar_status_vaga


# Gestão de Entrevistas
class AgendarEntrevistaView(LoginRequiredMixin, generic.UpdateView):
    """Agendar ou Reagendar Entrevista para um Candidato."""
    model = Entrevista
    form_class = EntrevistaForm
    template_name = 'candidaturas/entrevistas/form_agendar.html'
    
    def get_object(self, queryset=None):
        # Tenta pegar a entrevista existente ou cria uma nova instância não salva
        candidato_id = self.kwargs.get('pk')
        candidato = get_object_or_404(Candidato, pk=candidato_id)
        # Definir local padrão baseado na localização do candidato
        local_default = 'A definir'
        if candidato.distrito:
            local_default = f"STAE {candidato.distrito.nome}"
        elif candidato.provincia:
            local_default = f"STAE Província de {candidato.provincia.nome}"

        entrevista, pv = Entrevista.objects.get_or_create(
            candidato=candidato,
            defaults={
                'data_hora': timezone.now(),
                'local': local_default
            }
        )
        
        # Se já existia e estava com valor provisório, atualizar
        if not pv and entrevista.local == 'A definir':
            entrevista.local = local_default
            entrevista.save()
            
        return entrevista

    def form_valid(self, form):
        entrevista = form.save(commit=False)
        entrevista.status = Entrevista.Status.AGENDADA
        entrevista.save()
        
        # Atualizar estado do candidato
        candidato = entrevista.candidato
        candidato.estado = Candidato.Estado.ENTREVISTA_AGENDADA
        candidato.save()
        
        # Enviar notificacao (simulada por toast)
        messages.success(self.request, f"Entrevista agendada para {candidato.nome_completo} em {entrevista.data_hora}.")
        return redirect('candidaturas:detalhe_candidato', pk=candidato.pk)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        candidato_id = self.kwargs.get('pk')
        context['candidato'] = get_object_or_404(Candidato, pk=candidato_id)
        return context


class MinhasEntrevistasView(LoginRequiredMixin, generic.ListView):
    """Lista de entrevistas agendadas para o utilizador logado."""
    model = Entrevista
    template_name = 'candidaturas/entrevistas/minhas_entrevistas.html'
    context_object_name = 'entrevistas'
    
    def get_queryset(self):
        # Retorna entrevistas futuras onde o usuário é o entrevistador
        return Entrevista.objects.filter(
            entrevistador=self.request.user, 
            status=Entrevista.Status.AGENDADA
        ).order_by('data_hora')


class RealizarEntrevistaView(LoginRequiredMixin, generic.UpdateView):
    """Realizar a entrevista (Avaliação)."""
    model = Entrevista
    form_class = AvaliacaoEntrevistaForm
    template_name = 'candidaturas/entrevistas/form_realizar.html'
    
    def form_valid(self, form):
        entrevista = form.save(commit=False)
        entrevista.status = Entrevista.Status.REALIZADA
        entrevista.save()
        
        # Atualizar estado do candidato baseado no resultado
        candidato = entrevista.candidato
        if entrevista.resultado == Entrevista.Resultado.APROVADO:
            candidato.estado = Candidato.Estado.ENTREVISTA_APROVADA
            messages.success(self.request, "Candidato APROVADO na entrevista!")
        elif entrevista.resultado == Entrevista.Resultado.REPROVADO:
            candidato.estado = Candidato.Estado.ENTREVISTA_REPROVADA
            messages.warning(self.request, "Candidato REPROVADO na entrevista.")
        
        candidato.save()
        return redirect('candidaturas:minhas_entrevistas')

@login_required
def enviar_aprovados_lista_formacao(request):
    """
    Envia em massa os candidatos da vista de Aprovados para o DEFC.
    """
    if request.method != 'POST':
        return redirect('candidaturas:lista_verificacao')
        
    vaga_id = request.POST.get('vaga')
    genero = request.POST.get('genero')
    provincia_id = request.POST.get('provincia')
    
    # 1. Obter a base acessível
    qs = Candidato.objects.filter(estado=Candidato.Estado.ENTREVISTA_APROVADA)
    
    # Check permissions based on same logic from ListaVerificacaoView
    perfil = obter_perfil_usuario(request.user)
    if not request.user.is_superuser:
        if perfil.nivel == PerfilUtilizador.Nivel.PROVINCIAL and perfil.provincia:
            qs = qs.filter(provincia=perfil.provincia)
        elif perfil.nivel == PerfilUtilizador.Nivel.DISTRITAL and perfil.distrito:
            qs = qs.filter(distrito=perfil.distrito)

    # 2. Aplicar os filtros UI
    if vaga_id:
        qs = qs.filter(vaga__id=vaga_id)
    if genero:
        qs = qs.filter(genero=genero)
    if provincia_id:
        if request.user.is_superuser or perfil.nivel == PerfilUtilizador.Nivel.CENTRAL:
            qs = qs.filter(provincia__id=provincia_id)
            
    candidatos = list(qs)
    total = len(candidatos)
    
    if total == 0:
        messages.warning(request, "Nenhum candidato apto encontrado com os filtros atuais.")
        return redirect(f"{reverse('candidaturas:lista_verificacao')}?estado=ENTREVISTA_APROVADA")

    import random
    import string
    
    for c in candidatos:
         chars = string.ascii_letters + string.digits
         senha = ''.join(random.choices(chars, k=8))
         c.password_defc = senha
         c.estado = Candidato.Estado.ENVIADO_DEFC
         c.save()
         c.save()
         
    # Trigger DEFC migration script automatically
    import subprocess
    import platform
    import os
    python_cmd = 'python' if platform.system() == 'Windows' else 'python3'
    defc_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '../../sicfaae-defc'))
    script_path = os.path.join(defc_dir, 'migrate_defc.py')
    
    if os.path.exists(script_path):
        try:
            subprocess.Popen([python_cmd, script_path], cwd=defc_dir)
            messages.success(request, f"{total} candidatos aprovados foram enviados com sucesso e a sincronização com DEFC foi iniciada.")
        except Exception as e:
            messages.warning(request, f"{total} candidatos enviados, mas falha ao sincronizar com DEFC: {str(e)}")
    else:
        messages.success(request, f"{total} candidatos aprovados foram enviados para Formação com sucesso! (Sync script não encontrado)")
    
    # Base redirect with state
    url = f"{reverse('candidaturas:lista_verificacao')}?estado=ENVIADO_DEFC"
    if vaga_id: url += f"&vaga={vaga_id}"
    if genero: url += f"&genero={genero}"
    if provincia_id: url += f"&provincia={provincia_id}"
        
    return redirect(url)
